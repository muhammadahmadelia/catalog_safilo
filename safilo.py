import os
import re
import sys
import json
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from datetime import datetime

from models.store import Store
from models.product import Product
from models.variant import Variant
from models.metafields import Metafields
import glob
import requests

import threading
from openpyxl import Workbook
from openpyxl.drawing.image import Image as Imag
from PIL import Image
from lxml import html

from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

import warnings
warnings.filterwarnings("ignore")

class myScrapingThread(threading.Thread):
    def __init__(self, threadID: int, name: str, obj, brand_name: str, glasses_type: str, product_url: str, product_variations: list, cookies: dict, csrf_token: str) -> None:
        threading.Thread.__init__(self)
        self.threadID = threadID
        self.name = name
        self.brand_name = brand_name
        self.glasses_type = glasses_type
        self.product_url = product_url
        self.product_variations = product_variations
        self.cookies = cookies
        self.csrf_token = csrf_token
        self.obj = obj
        self.status = 'in progress'
        pass

    def run(self):
        self.obj.get_product_details(self.brand_name, self.glasses_type, self.product_url, self.product_variations, self.cookies, self.csrf_token)
        self.status = 'completed'

    def active_threads(self):
        return threading.activeCount()

class Safilo_Scraper:
    def __init__(self, DEBUG: bool, result_filename: str, logs_filename: str, chrome_path: str) -> None:
        self.DEBUG = DEBUG
        self.result_filename = result_filename
        self.logs_filename = logs_filename
        self.thread_list = []
        self.thread_counter = 0
        self.chrome_options = Options()
        self.chrome_options.add_argument('--disable-infobars')
        self.chrome_options.add_argument("--start-maximized")
        self.chrome_options.add_experimental_option('excludeSwitches', ['enable-logging'])
        self.args = ["hide_console", ]
        # self.browser = webdriver.Chrome(options=self.chrome_options, service_args=self.args)
        # self.browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=self.chrome_options)
        self.browser = webdriver.Chrome(service=ChromeService(chrome_path), options=self.chrome_options)
        self.data = []
        # self.ref_json_data = None
        pass

    def controller(self, store: Store, brands_with_types: list[dict]):
        try:

            
            brands_data: list = []
            cookies: dict = dict()

            self.browser.get(store.link)
            self.wait_until_browsing()
            self.accept_cookies()

            if self.login(store.username, store.password):

                print('Scraping products for')
                for brand_with_type in brands_with_types:
                    brand_name: str = brand_with_type['brand']
                    brand_code: str = brand_with_type['code']
                    # print(f'Brand: {brand_name}')

                    for glasses_type in brand_with_type['glasses_type']:
                        
                        print(f'Brand: {brand_name}')
                        self.print_logs(f'Brand: {brand_name}')

                        if not cookies: cookies = self.get_cookies()
                        if not brands_data: brands_data = self.get_brands_data(cookies)

                        if brands_data:
                            brand_json = self.get_brand_json(brand_name, brands_data, cookies)
                            if brand_json:
                                brand_url = brand_json['brand_url']
                                brand_category_id = brand_json['category_id']

                                self.open_new_tab(brand_url)
                                
                                csrf_token = self.get_csrf_token()
                                
                                data = self.get_all_product(brand_url, brand_category_id, glasses_type, cookies, csrf_token)
                                
                                total_products = data['total_products']
                                scraped_products = 0
                                start_time = datetime.now()

                                print(f'Type: {glasses_type} | Total products: {total_products}')
                                print(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                                self.print_logs(f'Type: {glasses_type} | Total products: {total_products}')
                                self.print_logs(f'Start Time: {start_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')

                                if total_products and int(total_products) > 0: 
                                    self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)
                                

                                for product_data in data['products']:
                                    product_url = product_data['url']
                                    product_variations = product_data['variations']
                                    
                                    # self.get_product_details(brand_name, glasses_type, product_url, product_variations, cookies, csrf_token)
                                    self.create_thread(brand_name, glasses_type, product_url, product_variations, cookies, csrf_token)
                                    if self.thread_counter >= 10: 
                                        self.wait_for_thread_list_to_complete()
                                        self.save_to_json(self.data)

                                    scraped_products += 1

                                    if total_products and int(total_products) > 0: 
                                        self.printProgressBar(scraped_products, total_products, prefix = 'Progress:', suffix = 'Complete', length = 50)

                                    self.save_to_json(self.data)

                                end_time = datetime.now()

                                print(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                print('Duration: {}\n'.format(end_time - start_time))

                                self.print_logs(f'End Time: {end_time.strftime("%A, %d %b %Y %I:%M:%S %p")}')
                                self.print_logs('Duration: {}\n'.format(end_time - start_time))
                                self.close_last_tab()

                        self.wait_for_thread_list_to_complete()
                        self.save_to_json(self.data)

                    self.wait_for_thread_list_to_complete()
                    self.save_to_json(self.data)
            else: print(f'Failed to login \nURL: {store.link}\nUsername: {str(store.username)}\nPassword: {str(store.password)}')
        except Exception as e:
            if self.DEBUG: print(f'Exception in scraper controller: {e}')
            else: pass
        finally: 
            self.wait_for_thread_list_to_complete()
            self.save_to_json(self.data)
            self.browser.quit()

    def wait_until_browsing(self) -> None:
        while True:
            try:
                state = self.browser.execute_script('return document.readyState; ')
                if 'complete' == state: break
                else: sleep(0.2)
            except: pass

    def wait_until_element_found(self, wait_value: int, type: str, value: str) -> bool:
        flag = False
        try:
            if type == 'id':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.ID, value)))
                flag = True
            elif type == 'xpath':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.XPATH, value)))
                flag = True
            elif type == 'css_selector':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CSS_SELECTOR, value)))
                flag = True
            elif type == 'class_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.CLASS_NAME, value)))
                flag = True
            elif type == 'tag_name':
                WebDriverWait(self.browser, wait_value).until(EC.presence_of_element_located((By.TAG_NAME, value)))
                flag = True
        except: pass
        finally: return flag

    def accept_cookies(self) -> None:
        try:
            # accept cookies if found
            if self.wait_until_element_found(30, 'xpath', '//button[contains(@id, "acceptCookiesPolicy")]'):
                for _ in range(0, 20):
                    try:
                        self.browser.find_element(By.XPATH,'//button[contains(@id, "acceptCookiesPolicy")]').click()
                        sleep(0.2)
                        break
                    except: sleep(0.5)
        except Exception as e:
            self.print_logs(f'Exception in accept_cookies: {str(e)}')
            if self.DEBUG: print(f'Exception in accept_cookies: {str(e)}')
            else: pass

    def login(self, email: str, password: str) -> bool:
        login_flag = False
        try:
            if self.wait_until_element_found(20, 'xpath', '//input[@class="username"]'):
                self.browser.find_element(By.XPATH, '//input[@class="username"]').send_keys(email)
                sleep(0.2)
                if self.wait_until_element_found(20, 'xpath', '//input[@class="password"]'):
                    self.browser.find_element(By.XPATH, '//input[@class="password"]').send_keys(password)
                    sleep(0.2)
                    self.browser.find_element(By.XPATH, '//button[@class="login-btn"]').click()

                    if self.wait_until_element_found(100, 'xpath', '//button/span[contains(text(), "Brands")]'): login_flag = True
                else: print('Password input not found')
            else: print('Email input not found')
        except Exception as e:
            self.print_logs(f'Exception in login: {str(e)}')
            if self.DEBUG: print(f'Exception in login: {str(e)}')
            else: pass
        finally: return login_flag
    
    def wait_for_page_loading(self):
        self.wait_until_browsing()
        for _ in range(0, 100):
            try:
                self.browser.find_element(By.XPATH, '//div[@id="overlay"]')
                sleep(0.5)
            except: break

    def get_headers(self, referer: str) -> dict:
        return {
            'authority': 'safilo.my.site.com',
            'accept': '*/*',
            'accept-language': 'en-US,en;q=0.9',
            'referer': referer,
            'sec-ch-ua': '"Not A(Brand";v="99", "Google Chrome";v="121", "Chromium";v="121"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'empty',
            'sec-fetch-mode': 'cors',
            'sec-fetch-site': 'same-origin',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36',
        }

    def get_brands_data(self, cookies) -> list:
        brands_data: list = []
        try:
            params = {
                'cacheable': 'true',
                'classname': '@udd/01p7T000000Clzf',
                'isContinuation': 'false',
                'method': 'fetchMenuItems',
                'namespace': '',
                'params': '{"effectiveAccountId":null,"language":"en-US"}',
                'language': 'en-US',
                'asGuest': 'false',
                'htmlEncode': 'false',
            }
            headers = self.get_headers('https://safilo.my.site.com/safilob2b/')
            
            response = requests.get(
                'https://safilo.my.site.com/safilob2b/webruntime/api/apex/execute',
                params=params,
                cookies=cookies,
                headers=headers,
                verify=False
            )
            if response.status_code == 200:
                for returnValue in response.json()['returnValue']:
                    if returnValue['displayName'] == 'Brands':
                        brands_data = returnValue['childMenuItems']
                        break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brands_data: {e}')
            else: pass
        finally: return brands_data

    def get_cookies(self) -> dict:
        cookies: dict = {}
        try:
            for browser_cookie in self.browser.get_cookies():
                cookies[browser_cookie['name']] = browser_cookie['value']
            if self.DEBUG: print(f'Exception in get_cookies: {e}')
            self.print_logs(f'Exception in get_cookies: {e}')
        finally: return cookies

    def get_brand_json(self, brand_name: str, brands_data: list, cookies: dict) -> dict:
        brand_json: dict = {}
        try:
            for childMenuItem in brands_data:
                if  str(brand_name).strip().lower() == str(childMenuItem['displayName']).strip().lower():
                    category_id = childMenuItem.get('item').get('LEX_Category__c')

                    headers = self.get_headers('https://safilo.my.site.com/safilob2b/')

                    API = f'https://safilo.my.site.com/safilob2b/webruntime/api/services/data/v59.0/connect/communities/0DB7T000000XZCrWAO/seo/properties/{category_id}?language=en-US&asGuest=false&htmlEncode=false'
                    response = requests.get(
                        url=API,
                        cookies=cookies,
                        headers=headers,
                        verify=False
                    )
                    if response.status_code == 200:
                        brand_url = response.json()['canonicalUrl']
                        brand_json = {
                            'category_id': category_id,
                            'brand_url': brand_url
                        }
                        break
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_brand_json: {e}')
            else: pass
        finally: return brand_json

    def get_csrf_token(self) -> str:
        csrf_token: str = ''
        try:
            self.wait_until_element_found(40, 'xpath', '//script[contains(text(), "csrfToken") and contains(text(), "isGuest") and contains(text(), "LWR.")]')
            doc_tree = html.fromstring(self.browser.page_source)
            text = doc_tree.xpath('//script[contains(text(), "csrfToken") and contains(text(), "isGuest") and contains(text(), "LWR.")]/text()')
            if text:
                csrf_token_match = re.search(r'"csrfToken":"(.*?)"', text[0])

                if csrf_token_match:
                    csrf_token = csrf_token_match.group(1)
                    csrf_token = bytes(csrf_token, "utf-8").decode("unicode_escape")
            else:
                input('No text found for csrf')
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_csrf_token: {e}')
            else: pass
        finally: return csrf_token

    def get_all_product(self, brand_url: str, brand_category_id:str, glasses_type, cookies: dict, csrf_token: str) -> dict:
        product_urls = []
        total_products = 0
        try:
            page_no = 0
            scraped_products = 0
            product_urls, total_products = self.get_product_urls(brand_url, brand_category_id, page_no, glasses_type, cookies, csrf_token)
            if int(total_products) > 0:
                scraped_products = len(product_urls)
                while scraped_products < total_products:
                    # print(page_no, scraped_products, total_products)
                    page_no += 1
                    new_product_urls, total_products = self.get_product_urls(brand_url, brand_category_id, page_no, glasses_type, cookies, csrf_token)

                    product_urls += new_product_urls

                    scraped_products = len(product_urls)
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_all_product: {e}')
            else: pass
        finally: return { 'products': product_urls, 'total_products': total_products }

    def get_product_urls(self, brand_url: str, category_id:str,  page: int, glasses_type, cookies: dict, csrf_token: str) -> dict:
        data = []
        total_products = 0
        try:
            product_urls = []
            
            # data ='{"namespace":"","classname":"@udd/01p7T000000Cm2N","method":"productSearch","isContinuation":false,"params":{"communityId":"0DB7T000000XZCrWAO","searchQuery":"{\\"searchTerm\\":\\"\\",\\"categoryId\\":\\"'+str(category_id)+'\\",\\"refinements\\":[{\\"nameOrId\\":\\"B2B_Adv__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_New__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_ProductTypeSpec__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[\\"'+str(glasses_type)+'\\"]},{\\"nameOrId\\":\\"B2B_SegmentSpec__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_FrameMaterial__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_LensWidthSize__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_ColorFamily__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_RimType__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_FrameShape__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]},{\\"nameOrId\\":\\"B2B_LensesDiagonal__c\\",\\"type\\":\\"DistinctValue\\",\\"attributeType\\":\\"Custom\\",\\"values\\":[]}],\\"page\\":'+str(page)+',\\"includePrices\\":true,\\"parentFields\\":[\\"StockKeepingUnit\\",\\"B2B_Brand__c\\",\\"B2B_OldName__c\\",\\"B2B_RestylingYear__c\\",\\"B2B_BioBasedMaterial__c\\",\\"B2B_LowChemicals__c\\",\\"B2B_Recycled__c\\",\\"B2B_BpaFree__c\\",\\"SFA_SortingOrder__c\\",\\"B2B_Adv__c\\",\\"DisplayUrl\\",\\"B2B_CoreCollection__c\\",\\"B2B_New__c\\"],\\"sortRuleId\\":null}","effectiveAccountId":null},"cacheable":false}'
            json_data = {
                'namespace': '',
                'classname': '@udd/01p7T000000Cm2N',
                'method': 'productSearch',
                'isContinuation': False,
                'params': {
                    'communityId': '0DB7T000000XZCrWAO',
                    'searchQuery': '{"searchTerm":"","categoryId":"'+str(category_id)+'","refinements":[{"nameOrId":"B2B_Adv__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_New__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_ProductTypeSpec__c","type":"DistinctValue","attributeType":"Custom","values":["'+str(glasses_type)+'"]},{"nameOrId":"B2B_SegmentSpec__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_FrameMaterial__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_LensWidthSize__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_ColorFamily__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_RimType__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_FrameShape__c","type":"DistinctValue","attributeType":"Custom","values":[]},{"nameOrId":"B2B_LensesDiagonal__c","type":"DistinctValue","attributeType":"Custom","values":[]}],"page":'+str(page)+',"includePrices":true,"parentFields":["StockKeepingUnit","B2B_Brand__c","B2B_OldName__c","B2B_RestylingYear__c","B2B_BioBasedMaterial__c","B2B_LowChemicals__c","B2B_Recycled__c","B2B_BpaFree__c","SFA_SortingOrder__c","B2B_Adv__c","DisplayUrl","B2B_CoreCollection__c","B2B_New__c"],"sortRuleId":null}',
                    'effectiveAccountId': None,
                },
                'cacheable': False,
            }
            API = 'https://www.youandsafilo.com/webruntime/api/apex/execute?language=en-US&asGuest=false&htmlEncode=false'
            headers = self.get_headers(brand_url)

            headers['csrf-token'] = csrf_token
            headers['origin'] = 'https://safilo.my.site.com'

            response = requests.post(
                url=API,
                cookies=cookies,
                headers=headers,
                json=json_data,
                verify=False
            )
            
            if response.status_code == 200:
                total_products = response.json()['returnValue']['productsPage']['total']

                for product_json in response.json()['returnValue']['productsPage']['products']:
                    id = product_json['id']
                    name = str(product_json['name']).strip().lower().replace(' ', '-').replace('/', '')
                    product_url = f'https://safilo.my.site.com/safilob2b/product/{name}/{id}'
                    
                    variations = []
                    for variationData in product_json['variationData']:
                        if 'prices' in variationData:
                            if variationData.get('prices').get('productId') not in variations:
                                variations.append(variationData.get('prices').get('productId'))

                    if product_url not in product_urls:
                        data.append({'url': product_url, 'variations': variations})
        
        except Exception as e:
            if self.DEBUG: print(f'Exception in get_product_urls: {e}')
            else: pass
        finally: return data, total_products

    def get_product_details(self, brand_name: str, glasses_type: str, product_url: str, variationIds: list, cookies: dict, csrf_token: str):
        try:
            headers = self.get_headers(product_url)

            headers['csrf-token'] = csrf_token
            headers['origin'] = 'https://safilo.my.site.com'

            variationIds = json.dumps(variationIds).replace('"', '\\"')
            modelId = str(product_url).split('/')[-1].strip()
            json_data = {
                'namespace': '',
                'classname': '@udd/01p7T000000Cm1Q',
                'method': 'getProductData',
                'isContinuation': False,
                'params': {
                    'inputDataString': '{"modelId":"'+str(modelId)+'","variationIds":"'+variationIds+'","communityId":"0DB7T000000XZCrWAO","effectiveAccountId":null}',
                },
                'cacheable': False,
            }

            API = 'https://www.youandsafilo.com/webruntime/api/apex/execute?language=en-US&asGuest=false&htmlEncode=false'

            response = requests.post(
                url=API,
                cookies=cookies,
                headers=headers,
                json=json_data,
                verify=False
            )
            if response.status_code == 200:
                json_data = response.json()['returnValue']
                
                product_number, product_name = '', ''
                try: product_number = json_data['model']['StockKeepingUnit']
                except: pass
                try: 
                    product_name = str(json_data['model']['Name']).strip().lower().replace(str(brand_name).strip().lower(), '').strip().upper()
                    product_name = self.clean_product_name(product_name)
                except: pass
                
                frame_codes = []

                for key, value in json_data['variationIdToVariations'].items():
                    frame_code = ''
                    try: frame_code = value['B2B_ColorCode__c']
                    except: pass

                    if frame_code not in frame_codes:
                        frame_codes.append(frame_code)

                        product = Product()
                        product.number = product_number
                        product.name = product_name
                        product.brand = brand_name
                        product.type = glasses_type
                        product.url = product_url

                        try: product.frame_code = value['B2B_ColorCode__c']
                        except: pass
                        try: product.lens_code = value['B2B_LensCode__c']
                        except: pass
                        try: product.bridge = value['B2B_BridgeLengthSize__c']
                        except: pass
                        try: product.template = str(int(value['B2B_TempleLengthSize__c']))
                        except: pass
                        try: product.image = str(value['DisplayUrl']).strip().replace('{0}', '00').replace('{1}', 'medium')
                        except: pass

                        # print(frame_code, lens_code, bridge, template, image_url)
                        metafields = Metafields()
                        try: metafields.for_who = value['segmentspec']
                        except: pass
                        try: metafields.lens_material = value['B2B_LensesMaterial__c']
                        except: pass
                        try: metafields.frame_shape = value['frameshape']
                        except: pass
                        try: metafields.frame_material = value['framematerial']
                        except: pass
                        try: metafields.frame_color = value['colorfamily'] if value['colorfamily'] else value['LEX_DescriptionRT__c']
                        except: pass
                        # print(gender, lens_material, frame_shape, frame_material, frame_color)

                        product.metafields = metafields

                        variant = Variant()
                        try: variant.title = value['B2B_LensWidthSize__c']
                        except: pass
                        try: variant.sku = value['Name'] if value['Name'] else value['StockKeepingUnit']
                        except: pass
                        try: variant.inventory_quantity = 5 if value['B2B_StockValue__c'] > 0 else 0
                        except: pass
                        try: variant.listing_price = self.get_price(json_data['pricebookEntriesByIds'], key)
                        except: pass
                        try: variant.barcode_or_gtin = value['B2B_EANCode__c']
                        except: pass
                        product.add_single_variant(variant)

                        self.data.append(product)
                    else:
                        for product in self.data:
                            if product.number == product_number and product.frame_code == frame_code:
                                variant = Variant()
                                try: variant.title = value['B2B_LensWidthSize__c']
                                except: pass
                                try: variant.sku = value['Name'] if value['Name'] else value['StockKeepingUnit']
                                except: pass
                                try: variant.inventory_quantity = 5 if value['B2B_StockValue__c'] > 0 else 0
                                except: pass
                                try: variant.listing_price = self.get_price(json_data['pricebookEntriesByIds'], key)
                                except: pass
                                try: variant.barcode_or_gtin = value['B2B_EANCode__c']
                                except: pass
                                product.add_single_variant(variant)
                                break

        except Exception as e:
            if self.DEBUG: print(f'Exception in get_product_details: {e}')
            else: pass
    
    def clean_product_name(self, product_name: str) -> None:
        try:
            if 'CA' in product_name.split(' '): product_name = product_name.replace('CA ', '').strip()
            elif 'CARDUC' in product_name.split(' '): product_name = product_name.replace('CARDUC ', '').strip()
            elif 'CF' in product_name.split(' '): product_name = product_name.replace('CF ', '').strip()
            elif 'DB' in product_name.split(' '): product_name = product_name.replace('DB ', '').strip()
            elif 'PLD' in product_name.split(' '): product_name = product_name.replace('PLD ', '').strip()
            elif 'MARC' in product_name.split(' '): product_name = product_name.replace('MARC ', '').strip()
            elif 'MJ' in product_name.split(' '): product_name = product_name.replace('MJ ', '').strip()
        except Exception as e:
            if self.DEBUG: print(f'Exception in clean_product_name: {e}')
            self.print_logs(f'Exception in clean_product_name: {e}')
        finally: return product_name

    def get_price(self, data, varation_key):
        for key, value in data.items():
            if value['productId'] == varation_key:
                return value['retailPrice']

    def open_new_tab(self, url: str) -> None:
        # open category in new tab
        self.browser.execute_script('window.open("'+str(url)+'","_blank");')
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])
        self.wait_until_browsing()

    def close_last_tab(self) -> None:
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[len(self.browser.window_handles) - 1])

    def create_thread(self, brand_name: str, glasses_type: str, product_url: str, product_variations: list, cookies: dict, csrf_token: str):
        thread_name = "Thread-"+str(self.thread_counter)
        self.thread_list.append(myScrapingThread(self.thread_counter, thread_name, self, brand_name, glasses_type, product_url, product_variations, cookies, csrf_token))
        self.thread_list[self.thread_counter].start()
        self.thread_counter += 1

    def is_thread_list_complted(self):
        for obj in self.thread_list:
            if obj.status == "in progress":
                return False
        return True

    def wait_for_thread_list_to_complete(self):
        while True:
            result = self.is_thread_list_complted()
            if result: 
                self.thread_counter = 0
                self.thread_list.clear()
                break
            else: sleep(1)

    def save_to_json(self, products: list[Product]):
        try:
            json_products = []
            for product in products:
                json_varinats = []
                for index, variant in enumerate(product.variants):
                    json_varinat = {
                        'position': (index + 1), 
                        'title': variant.title, 
                        'sku': variant.sku, 
                        'inventory_quantity': variant.inventory_quantity,
                        'found_status': variant.found_status,
                        'wholesale_price': variant.wholesale_price,
                        'listing_price': variant.listing_price, 
                        'barcode_or_gtin': variant.barcode_or_gtin,
                        'size': variant.size,
                    }
                    json_varinats.append(json_varinat)
                json_product = {
                    'brand': product.brand, 
                    'number': product.number, 
                    'name': product.name, 
                    'frame_code': product.frame_code,  
                    'lens_code': product.lens_code, 
                    'lens_color': product.metafields.lens_color,
                    'frame_color': product.metafields.frame_color,
                    # 'status': product.status, 
                    # 'type': product.type, 
                    'url': product.url, 
                    'metafields': [
                        { 'key': 'for_who', 'value': product.metafields.for_who },
                        # { 'key': 'product_size', 'value': product.metafields.product_size }, 
                        { 'key': 'lens_material', 'value': product.metafields.lens_material }, 
                        { 'key': 'lens_technology', 'value': product.metafields.lens_technology }, 
                        { 'key': 'frame_material', 'value': product.metafields.frame_material }, 
                        { 'key': 'frame_shape', 'value': product.metafields.frame_shape },
                        { 'key': 'gtin1', 'value': product.metafields.gtin1 }, 
                        { 'key': 'img_url', 'value': product.image }
                    ],
                    'variants': json_varinats
                }
                json_products.append(json_product)
            
           
            with open(self.result_filename, 'w') as f: json.dump(json_products, f)
            
        except Exception as e:
            if self.DEBUG: print(f'Exception in save_to_json: {e}')
            else: pass

    # print logs to the log file
    def print_logs(self, log: str) -> None:
        try:
            with open(self.logs_filename, 'a') as f:
                f.write(f'\n{log}')
        except: pass

    def printProgressBar(self, iteration, total, prefix = '', suffix = '', decimals = 1, length = 100, fill = '█', printEnd = "\r") -> None:
        """
        Call in a loop to create terminal progress bar
        @params:
            iteration   - Required  : current iteration (Int)
            total       - Required  : total iterations (Int)
            prefix      - Optional  : prefix string (Str)
            suffix      - Optional  : suffix string (Str)
            decimals    - Optional  : positive number of decimals in percent complete (Int)
            length      - Optional  : character length of bar (Int)
            fill        - Optional  : bar fill character (Str)
            printEnd    - Optional  : end character (e.g. "\r", "\r\n") (Str)
        """
        percent = ("{0:." + str(decimals) + "f}").format(100 * (iteration / float(total)))
        filledLength = int(length * iteration // total)
        bar = fill * filledLength + '-' * (length - filledLength)
        print(f'\r{prefix} |{bar}| {percent}% {suffix}', end = printEnd)
        # Print New Line on Complete
        if iteration == total: 
            print()

def read_data_from_json_file(DEBUG, result_filename: str):
    data = []
    try:
        files = glob.glob(result_filename)
        if files:
            f = open(files[-1])
            json_data = json.loads(f.read())
            products = []

            for json_d in json_data:
                number, frame_code, brand, img_url, frame_color, lens_color = '', '', '', '', '', ''
                # product = Product()
                brand = json_d['brand']
                number = str(json_d['name']).strip().upper()
                if '-' in number: number = number.replace('-', '/').strip()
                # product.name = str(json_d['name']).strip().upper()
                frame_code = str(json_d['frame_code']).strip().upper()
                if '-' in frame_code: frame_code = frame_code.replace('-', '/').strip()
                frame_color = str(json_d['frame_color']).strip().title()
                # product.lens_code = str(json_d['lens_code']).strip().upper()
                lens_color = str(json_d['lens_color']).strip().title()
                # product.status = str(json_d['status']).strip().lower()
                # product.type = str(json_d['type']).strip().title()
                # product.url = str(json_d['url']).strip()
                # metafields = Metafields()
                
                for json_metafiels in json_d['metafields']:
                    # if json_metafiels['key'] == 'for_who':metafields.for_who = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'product_size':metafields.product_size = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'activity':metafields.activity = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_material':metafields.lens_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'graduabile':metafields.graduabile = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'interest':metafields.interest = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'lens_technology':metafields.lens_technology = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_material':metafields.frame_material = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'frame_shape':metafields.frame_shape = str(json_metafiels['value']).strip().title()
                    # elif json_metafiels['key'] == 'gtin1':metafields.gtin1 = str(json_metafiels['value']).strip().title()
                    if json_metafiels['key'] == 'img_url':img_url = str(json_metafiels['value']).strip()
                    # elif json_metafiels['key'] == 'img_360_urls':
                    #     value = str(json_metafiels['value']).strip()
                    #     if '[' in value: value = str(value).replace('[', '').strip()
                    #     if ']' in value: value = str(value).replace(']', '').strip()
                    #     if "'" in value: value = str(value).replace("'", '').strip()
                    #     for v in value.split(','):
                    #         metafields.img_360_urls = str(v).strip()
                # product.metafields = metafields
                for json_variant in json_d['variants']:
                    sku, price = '', ''
                    # variant = Variant()
                    # variant.position = json_variant['position']
                    # variant.title = str(json_variant['title']).strip()
                    sku = str(json_variant['sku']).strip().upper()
                    if '/' in sku: sku = sku.replace('/', '-').strip()
                    # variant.inventory_quantity = json_variant['inventory_quantity']
                    # variant.found_status = json_variant['found_status']
                    wholesale_price = str(json_variant['wholesale_price']).strip()
                    listing_price = str(json_variant['listing_price']).strip()
                    barcode_or_gtin = str(json_variant['barcode_or_gtin']).strip()
                    # variant.size = str(json_variant['size']).strip()
                    # variant.weight = str(json_variant['weight']).strip()
                    # product.variants = variant
                    image_filname = f'Images/{sku}.jpg'
                    if not os.path.exists(image_filname):
                        image_attachment = download_image(img_url)
                        if image_attachment:
                            with open(f'Images/{sku}.jpg', 'wb') as f: f.write(image_attachment)
                            crop_downloaded_image(f'Images/{sku}.jpg')

                    data.append([brand, number, frame_code, frame_color, lens_color,  sku, wholesale_price, listing_price, barcode_or_gtin])
    except Exception as e:
        if DEBUG: print(f'Exception in read_data_from_json_file: {e}')
        else: pass
    finally: return data

def download_image(url):
    image_attachment = ''
    try:
        headers = {
            'accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
            'accept-Encoding': 'gzip, deflate, br',
            'accept-Language': 'en-US,en;q=0.9',
            'cache-Control': 'max-age=0',
            'sec-ch-ua': '"Google Chrome";v="95", "Chromium";v="95", ";Not A Brand";v="99"',
            'sec-ch-ua-mobile': '?0',
            'sec-ch-ua-platform': '"Windows"',
            'sec-fetch-dest': 'document',
            'sec-fetch-mode': 'navigate',
            'sec-fetch-site': 'none',
            'Sec-Fetch-User': '?1',
            'upgrade-insecure-requests': '1',
        }
        counter = 0
        while True:
            try:
                response = requests.get(url=url, headers=headers, timeout=20)
                # print(response.status_code)
                if response.status_code == 200:
                    # image_attachment = base64.b64encode(response.content)
                    image_attachment = response.content
                    break
                else: print(f'{response.status_code} found for downloading image')
            except: sleep(0.3)
            counter += 1
            if counter == 10: break
    except Exception as e: print(f'Exception in download_image: {str(e)}')
    finally: return image_attachment

def crop_downloaded_image(filename):
    try:
        im = Image.open(filename)
        width, height = im.size   # Get dimensions
        new_width = 1120
        new_height = 600
        if width > new_width and height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
        elif height > new_height:
            left = (width - new_width)/2
            top = (height - new_height)/2
            right = (width + new_width)/2
            bottom = (height + new_height)/2
            im = im.crop((left, top, right, bottom))
            im.save(filename)
    except Exception as e: print(f'Exception in crop_downloaded_image: {e}')

def saving_picture_in_excel(data: list):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.cell(row=1, column=1, value='Brand')
    worksheet.cell(row=1, column=2, value='Model Code')
    worksheet.cell(row=1, column=3, value='Lens Code')
    worksheet.cell(row=1, column=4, value='Color Frame')
    worksheet.cell(row=1, column=5, value='Color Lens')
    worksheet.cell(row=1, column=6, value='SKU')
    worksheet.cell(row=1, column=7, value='Wholesale Price')
    worksheet.cell(row=1, column=8, value='Listing Price')
    worksheet.cell(row=1, column=9, value="UPC")
    worksheet.cell(row=1, column=10, value="Image")

    for index, d in enumerate(data):
        new_index = index + 2

        worksheet.cell(row=new_index, column=1, value=d[0])
        worksheet.cell(row=new_index, column=2, value=d[1])
        worksheet.cell(row=new_index, column=3, value=d[2])
        worksheet.cell(row=new_index, column=4, value=d[3])
        worksheet.cell(row=new_index, column=5, value=d[4])
        worksheet.cell(row=new_index, column=6, value=d[5])
        worksheet.cell(row=new_index, column=7, value=d[6])
        worksheet.cell(row=new_index, column=8, value=d[7])
        worksheet.cell(row=new_index, column=9, value=d[8])

        image = f'Images/{d[-4]}.jpg'
        if os.path.exists(image):
            im = Image.open(image)
            width, height = im.size
            worksheet.row_dimensions[new_index].height = height
            worksheet.add_image(Imag(image), anchor='J'+str(new_index))
            # col_letter = get_column_letter(7)
            # worksheet.column_dimensions[col_letter].width = width

    workbook.save('Safilo Results.xlsx')

DEBUG = True
try:
    pathofpyfolder = os.path.realpath(sys.argv[0])
    # get path of Exe folder
    path = pathofpyfolder.replace(pathofpyfolder.split('\\')[-1], '')
    
    if os.path.exists('Safilo Results.xlsx'): os.remove('Safilo Results.xlsx')

    if '.exe' in pathofpyfolder.split('\\')[-1]: DEBUG = False
    
    f = open('Safilo start.json')
    json_data = json.loads(f.read())
    f.close()

    brands = json_data['brands']

    
    f = open('requirements/safilo.json')
    data = json.loads(f.read())
    f.close()

    store = Store()
    store.link = data['url']
    store.username = data['username']
    store.password = data['password']
    store.login_flag = True

    result_filename = 'requirements/Safilo Results.json'

    if not os.path.exists('Logs'): os.makedirs('Logs')

    log_files = glob.glob('Logs/*.txt')
    if len(log_files) > 5:
        oldest_file = min(log_files, key=os.path.getctime)
        os.remove(oldest_file)
        log_files = glob.glob('Logs/*.txt')

    scrape_time = datetime.now().strftime('%d-%m-%Y %H-%M-%S')
    logs_filename = f'Logs/Logs {scrape_time}.txt'

    chrome_path = ''
    if not chrome_path:
        chrome_path = ChromeDriverManager().install()
        if 'chromedriver.exe' not in chrome_path:
            chrome_path = str(chrome_path).split('/')[0].strip()
            chrome_path = f'{chrome_path}\\chromedriver.exe'
    
    Safilo_Scraper(DEBUG, result_filename, logs_filename, chrome_path).controller(store, brands)
    
    for filename in glob.glob('Images/*'): os.remove(filename)
    data = read_data_from_json_file(DEBUG, result_filename)
    os.remove(result_filename)

    saving_picture_in_excel(data)
except Exception as e:
    if DEBUG: print('Exception: '+str(e))
    else: pass
